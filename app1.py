import openpyxl as xl
from openpyxl.chart import BarChart,Reference
wb=xl.load_workbook('C:\MeeraHarish\Transactions.xlsx')
sheet=wb['Sheet1']

for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)
    updated_price=0.9*cell.value
    updated_price_cell=sheet.cell(row,4)
    updated_price_cell.value=updated_price

values=Reference(sheet,
                 min_row=2,
                 max_row=sheet.max_row,
                 min_col=4,
                 max_col=4
                )

chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'G2')
wb.save('transactions3.xlsx')
